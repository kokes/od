import csv

from openpyxl import load_workbook


exphd = ['Evidenční číslo dotace',
         'Identifikator dotace',
         'Název dotace',
         'Účastník',
         'IČ účastníka',
         'Účel dotace',
         'Poskytovatel dotace',
         'IČ poskytovatele',
         'Částka požadovaná',
         'Částka schválená',
         'Datum poskytnutí dotace']

hdclean = ['evidencni_cislo_dotace',
           'identifikator_dotace',
           'nazev_dotace',
           'ucastnik',
           'ic_ucastnika',
           'ucel_dotace',
           'poskytovatel_dotace',
           'ic_poskytovatele',
           'castka_pozadovana',
           'castka_schvalne',
           'datum_poskytnuti']


if __name__ == '__main__':
    # TODO: NULL_x000d
    wb = load_workbook(
        'data/raw/DotInfo_report_13_07_2017.xlsx', read_only=True)
    ws = wb.active

    rows = ws.iter_rows()

    # hlavicka by mela souhlasit s definici vyse
    # v excelu bude o sloupec vice, ten nas nezajima
    first = next(rows)
    hd = [j.value for j in first]
    assert hd[:len(exphd)] == exphd

    with open('data/dotinfo.csv', 'w') as fw:
        cw = csv.writer(fw)
        cw.writerow(hdclean)

        for row in rows:
            dt = [None if j.value ==
                  'NULL' else j.value for j in row][:len(exphd)]
            if not (isinstance(dt[4] or 0, int) and isinstance(dt[7] or 0, int)):
                print('korumpovana radka', dt)  #  chyby v 7/2017 exportu
                continue
            cw.writerow(dt)
