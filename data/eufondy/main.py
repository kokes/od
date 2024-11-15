# 2007-2013:
# http://dotaceeu.cz/cs/Evropske-fondy-v-CR/Programove-obdobi-2007-2013/Cerpani-v-obdobi-2007-2013
# http://dotaceeu.cz/Dotace/media/SF/Informace%20o%20%c4%8derp%c3%a1n%c3%ad/Seznamy%20p%c5%99%c3%adjemc%c5%af%20(List%20of%20Beneficiaries)/2016/Seznam-prijemcu-05_2017.xlsx

# TODO: osetrit tady cesko-polsko veci
# TODO: přehled projektů pro 2007-2013? je tam víc info
import csv
import io
import json
import os
import re
from functools import partial
from itertools import zip_longest
from urllib.request import urlopen

from lxml.etree import iterparse
from openpyxl import load_workbook

MS_2014_URL = "https://ms14opendata.mssf.cz/SeznamProjektu.xml"
MS_2021_URL = "https://ms21opendata.mssf.cz/SeznamOperaci_21_27.xml"


# TODO: nechcem strptime?
def predatuj(s):
    if s is None or len(s) == 0:
        return None

    d, m, y = map(int, s.split("."))
    return f"{y}-{m:02d}-{d:02d}"


def find_el(el, path):
    ns = el.nsmap[None]
    parts = path.split("/")
    fel = el.find("{{{}}}".format(ns) + ("/{{{}}}".format(ns)).join(parts))
    return getattr(fel, "text", None)


def parse_el(el, paths):
    ret = {}
    for k, v in paths.items():
        if isinstance(v, dict):
            ret[k] = json.dumps(parse_el(el, v), ensure_ascii=False)
        else:
            ret[k] = find_el(el, v)

    return ret if len(ret) > 0 else None


schema_od = {
    "id": "ID",
    "id_vyzva": "ID_VYZVA",
    "kod": "KOD",
    "naz": "NAZ",
    "nazeva": "NAZEVA",
    "popis": "POPIS",
    "problem": "PROBLEM",
    "cil": "CIL",
    "datum_zahajeni": "DZRSKUT",
    "datum_ukonceni_predp": "DURPRED",
    "datum_ukonceni_skut": "DURSKUT",
    "suk": "SUK",
    "zadatel_nazev": "ZAD/NAZ",
    "zadatel_ico": "ZAD/IC",
    "zadatel_pravni_forma": "ZAD/HPF",
    "zadatel_adresa": {
        "ruian": "ZAD/ADR/RUIAN",
        "kkod": "ZAD/ADR/KKOD",
        "knazev": "ZAD/ADR/KNAZEV",
        "okkod": "ZAD/ADR/OKKOD",
        "oknazev": "ZAD/ADR/OKNAZEV",
        "obkod": "ZAD/ADR/OBKOD",
        "obnazev": "ZAD/ADR/OBNAZEV",
        "cobcenazev": "ZAD/ADR/COBCENAZEV",
        "psc": "ZAD/ADR/PSC",
        "cp": "ZAD/ADR/CP",
        "cisor": "ZAD/ADR/CISOR",
        "ul": "ZAD/ADR/UL",
        "www": "ZAD/ADR/WWW",
    },
    "cile_projektu": "PRJSC/SC",
    # 'um': 'UM', # TODO: rozlisit dopadove umisteni a realizacni
    "financovani_czv": "PF/CZV",
    "financovani_eu": "PF/EU",
    "financovani_cnv": "PF/CNV",
    "financovani_sn": "PF/SN",
    "financovani_s": "PF/S",
    "financovani_esif": "PF/ESIF",
    "financovani_cv": "PF/CV",
    "cilove_skupiny": "CILSKUP/CSKOD",
}


# TODO: refaktorovat vsechny `prehled_` funkce, aby cetly jednu JSON specku
# pouzit DictWriter vsude, jako tady
def prehled_2021_2027(outdir: str, partial: bool = False):
    cdir = os.path.dirname(os.path.abspath(__file__))
    with open(os.path.join(cdir, "hlavicka2127.json"), encoding="utf8") as f:
        hd = json.load(f)

    source_url = (
        "https://dotaceeu.cz/getmedia/23d032d3-55b4-4150-9349-55a015eb4ae0/"
        "2024_08_Seznam-operaci_List-of-Operations_21.xlsx.aspx?ext=.xlsx"
    )
    print(
        f"Stahuji z seznam operaci z {source_url}, ale nemusi to byt nejaktualnejsi "
        f"export - prekontroluj stranku https://dotaceeu.cz/cs/statistiky-a-analyzy/"
        f"seznamy-prijemcu"
    )
    with urlopen(source_url, timeout=60) as r:
        raw = io.BytesIO(r.read())

    wb = load_workbook(raw, read_only=True)
    sh = wb.active
    assert sh.title == "Sheet1"
    rows = sh.iter_rows()
    next(rows), next(rows)  # nadpis, datum generovani
    next(rows)  # prazdna radka

    fr = [j.value.strip() for j in next(rows) if j.value is not None]
    assert fr == hd["ocekavane"], [
        (a, b) for a, b in zip_longest(fr, hd["ocekavane"]) if a != b
    ]
    next(rows)  # anglicky nazvy

    with open(
        os.path.join(outdir, "prehled_2021_2027.csv"), "w", encoding="utf8"
    ) as fw:
        cw = csv.DictWriter(fw, lineterminator="\n", fieldnames=hd["hlavicka"])
        cw.writeheader()
        for j, rrow in enumerate(rows):
            if partial and j > 1000:
                break
            rowv = [rrow[j].value for j in range(len(hd["hlavicka"]))]
            row = dict(zip(hd["hlavicka"], rowv))

            cw.writerow(row)


def prehled_2014_2020(outdir: str, partial: bool = False):
    cdir = os.path.dirname(os.path.abspath(__file__))
    with open(os.path.join(cdir, "hlavicka1420.json"), encoding="utf8") as f:
        hd = json.load(f)

    source_url = (
        "https://dotaceeu.cz/getmedia/47884143-6367-4ab3-ae32-d4a45b9d5ad5/"
        "2021_09_Seznam-operaci-_-List-of-operations.xlsx.aspx?ext=.xlsx"
    )
    print(
        f"Stahuji z seznam operaci z {source_url}, ale nemusi to byt nejaktualnejsi "
        f"export - prekontroluj stranku https://dotaceeu.cz/cs/statistiky-a-analyzy/"
        f"seznamy-prijemcu"
    )
    with urlopen(source_url, timeout=60) as r:
        raw = io.BytesIO(r.read())

    wb = load_workbook(raw, read_only=True)
    sh = wb.active
    assert sh.title == "Seznam operací"
    rows = sh.iter_rows()
    next(rows), next(rows)  # nadpis, datum generovani

    fr = [j.value.strip() for j in next(rows) if j.value is not None]
    assert fr == hd["ocekavane"], [
        (a, b) for a, b in zip_longest(fr, hd["ocekavane"]) if a != b
    ]
    next(rows)  # anglicky nazvy

    with open(
        os.path.join(outdir, "prehled_2014_2020.csv"), "w", encoding="utf8"
    ) as fw:
        cw = csv.writer(fw, lineterminator="\n")
        cw.writerow(hd["hlavicka"])
        for j, rrow in enumerate(rows):
            if partial and j > 1000:
                break
            row = [rrow[j].value for j in range(len(hd["hlavicka"]))]

            for cl in [9, 10, 11]:
                row[cl] = predatuj(row[cl])

            cw.writerow(row)


def prehled_2007_2013(outdir: str, partial: bool = False):
    cws = re.compile(r"\s+")

    source_url = (
        "http://dotaceeu.cz/Dotace/media/SF/Informace%20o%20%c4%8derp%c3%a1n%c3%ad/"
        "Seznamy%20p%c5%99%c3%adjemc%c5%af%20(List%20of%20Beneficiaries)/2016/"
        "Seznam-prijemcu-05_2017.xlsx"
    )
    # load_workbook zamyka vsechny soubory, tak to musime udelat pres pamet
    # nemam z toho radost, ale...
    with urlopen(source_url, timeout=60) as r:
        raw = io.BytesIO(r.read())

    wb = load_workbook(raw, read_only=True)
    sh = wb.active

    with open(
        os.path.join(outdir, "prehled_2007_2013.csv"), "w", encoding="utf8"
    ) as fw:
        cw = csv.writer(fw, lineterminator="\n")
        hd = [
            "prijemce",
            "ico",
            "projekt",
            "operacni_program",
            "fond_eu",
            "datum_alokace",
            "castka_alokovana",
            "datum_platby",
            "castka_proplacena",
            "stav",
        ]
        cw.writerow(hd)
        for j, row in enumerate(sh.rows):
            if partial and j > 1000:
                break
            dt = [j.value for j in row]
            assert len(dt) == 10
            if j == 0:
                assert dt[0] == (
                    "LIST OF BENEFICIARIES \nSEZNAM " "PŘÍJEMCŮ PODPORY Z FONDŮ EU"
                )
            elif j == 6:
                assert dt == [
                    " Název příjemce",
                    "IČ",
                    "Název projektu",
                    "Operační \nprogram",
                    "Fond\nEU",
                    "Částka hrazená z fondů EU ",
                    None,
                    None,
                    None,
                    None,
                ]
            elif j == 7:
                assert dt == [
                    None,
                    None,
                    None,
                    None,
                    None,
                    "Datum alokace",
                    "Alokovaná částka",
                    "Datum průběžné platby",
                    "Celková částka proplacená od začátku projektu",
                    "Stav",
                ]
            elif all([j is None for j in dt]):
                continue  # predposledni radka
            elif dt[0] == "Sestava vytvořena IS MSC2007":
                break  # koncime

            if j < 8:
                continue

            dt = [
                cws.sub(" ", j.strip()) if isinstance(j, str) else j for j in dt
            ]  # cistime bile znaky

            # ICO
            if dt[1] is None:
                pass
            # OP ČR-Polsko má polský IČ, který potřebujem vyfiltrovat
            # bohužel nejsou nějak jednoznačně určené
            elif dt[3] == "OP PS ČR-Polsko" and not dt[1].isdigit():
                dt[1] = None
            else:
                dt[1] = int(dt[1])

            dt[5] = predatuj(dt[5])  # datum alokace
            dt[7] = predatuj(dt[7])  # prubezna platba

            cw.writerow(dt)


def opendata_xml(url: str, fn: str, outdir: str, partial: bool = False):
    sloupce = [
        "id",
        "id_vyzva",
        "kod",
        "naz",
        "nazeva",
        "popis",
        "problem",
        "cil",
        "datum_zahajeni",
        "datum_ukonceni_predp",
        "datum_ukonceni_skut",
        "suk",
        "zadatel_nazev",
        "zadatel_ico",
        "zadatel_pravni_forma",
        "zadatel_adresa",
        "cile_projektu",
        "financovani_czv",
        "financovani_eu",
        "financovani_cnv",
        "financovani_sn",
        "financovani_s",
        "financovani_esif",
        "financovani_cv",
        "cilove_skupiny",
    ]

    with open(os.path.join(outdir, fn), "w", encoding="utf8") as fw:
        cw = csv.DictWriter(fw, fieldnames=sloupce, lineterminator="\n")
        cw.writeheader()
        r = urlopen(url, timeout=300)
        et = iterparse(r)

        for j, (action, element) in enumerate(et):
            if partial and j > 1e4:
                break
            assert action == "end"
            if not element.tag.endswith("}PRJ"):
                continue
            projekt = parse_el(element, schema_od)

            cw.writerow(projekt)
            element.clear()


opendata_2014_2020 = partial(opendata_xml, MS_2014_URL, "2014_2020.csv")
opendata_2021_2027 = partial(opendata_xml, MS_2021_URL, "2021_2027.csv")


# neimplementujem `partial`, protoze tech dat stejne neni moc
def main(outdir: str, partial: bool = False):
    prehled_2007_2013(outdir, partial)
    prehled_2014_2020(outdir, partial)
    prehled_2021_2027(outdir, partial)
    od_outdir = os.path.join(outdir, "opendata")
    os.makedirs(od_outdir, exist_ok=True)
    opendata_2014_2020(od_outdir, partial)
    opendata_2021_2027(od_outdir, partial)


if __name__ == "__main__":
    main(".")
